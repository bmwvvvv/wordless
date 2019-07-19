import re
import sys
import random
import openpyxl as xls
from PySide2 import QtCore,QtWidgets,QtGui


class Word:
    def __init__(self, index, liter, partOfSpeech, chinese, occurs, shots, flag):
        self.index   = index
        self.liter   = liter
        self.pospch  = partOfSpeech
        self.chinese = chinese
        self.occurs  = occurs
        self.shots   = shots
        self.flag    = flag

class Lexicon:
    def __init__(self, path):
        self.filename = path
        self.wb = xls.load_workbook(path)
        self.ws = self.wb.active
        self.totalWordNum = self.ws.max_row

    def updateOccursOfWord(self, word):
       self.ws.cell(column=4, row=word.index, value=word.occurs)

    def updateShotsOfWord(self, word):
       self.ws.cell(column=5, row=word.index, value=word.shots)

    def getOneWord(self, row = 1):
        if row is 0:
            wordAttrs = list(self.ws.rows)[0]
        else:
            wordAttrs = list(self.ws.rows)[row - 1]

        word = Word(row,
                    wordAttrs[0].value,     # literal
                    wordAttrs[1].value,     # part of speech
                    wordAttrs[2].value,     # Chinese
                    wordAttrs[3].value,     # occurs
                    wordAttrs[4].value,     # shots
                    wordAttrs[5].value)     # flag

        word.occurs += 1
        self.updateOccursOfWord(word)    # guarantee the consistency of data

        return word
        
class MyWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.lex = Lexicon('words.xlsx') 
        self.wordInd = random.randint(1, self.lex.totalWordNum)
        self.word = self.lex.getOneWord(self.wordInd)

        self.initUI()

    def initUI(self):
        self.button = QtWidgets.QPushButton("Click me!")

        self.curWord = QtWidgets.QLabel('<font face="verdana" color="maroon" size="10"><b>'+self.word.liter+'</b></font>')
        self.curWord.setAlignment(QtCore.Qt.AlignCenter)
        self.curWord.setScaledContents(True)


        self.prevWord = QtWidgets.QLabel(self.word.liter + "\n" + self.word.chinese)
        self.prevWord.setAlignment(QtCore.Qt.AlignCenter)
        self.prevWord.setScaledContents(True)

        self.negImage = QtGui.QPixmap("Images/negative.png")
        self.negImage.scaledToHeight(60)
        self.negImage.scaledToWidth(60)

        self.posImage = QtGui.QPixmap("Images/sam.png")
        self.posImage.scaledToHeight(60)
        self.posImage.scaledToWidth(60)

        self.hintImage = QtWidgets.QLabel()
        self.hintImage.setAlignment(QtCore.Qt.AlignCenter)
        self.hintImage.setPixmap(self.negImage)


        self.chinese = QtWidgets.QLineEdit()
        self.chinese.returnPressed.connect(self.magic)

        self.layout = QtWidgets.QGridLayout()
        self.layout.setSpacing(5)
        self.layout.addWidget(self.prevWord, 0, 0, 1, 1)
        self.layout.addWidget(self.curWord,  1, 1, 1, 1)
        self.layout.addWidget(self.hintImage,2, 1, 1, 1)
        self.layout.addWidget(self.chinese,  3, 1, 1, 1)
        self.layout.addWidget(self.button,   4, 1, 1, 1)
        self.layout.setColumnStretch(0, 1)
        self.layout.setColumnStretch(1, 2)
        self.layout.setColumnStretch(2, 1)
        self.layout.setRowStretch(0, 1)
        self.layout.setRowStretch(1, 2)
        self.layout.setRowStretch(2, 1)
        self.layout.setRowStretch(3, 1)
        
        self.setLayout(self.layout)

        self.button.clicked.connect(self.magic)

    def refreshPrevWord(self):
        self.prevWord.setText("<font face=verdana size=20><b>"+self.word.liter+"<br /></b></font>"
                             +"<font face=verdana size=20><b>"+self.word.chinese+"</b></font>")

    def retrieveWord(self):
        self.wordInd = random.randint(1, self.lex.totalWordNum)    
        self.word = self.lex.getOneWord(self.wordInd)
        if round(pow(2, self.word.shots/self.word.occurs)) is 2:
            tmpWordInd = random.randint(1, self.lex.totalWordNum)    
            if self.wordInd == tmpWordInd:
                pass
            else:
                self.wordInd = random.randint(1, self.lex.totalWordNum)    
                self.word = self.lex.getOneWord(self.wordInd)

        self.curWord.setText('<h6><font face="verdana" color="maroon" size="10"><b>'+self.word.liter+'</b></font></h6>')
        self.hintImage.setPixmap(self.negImage)
        self.hintImage.update()

    def updateWord(self):
        self.refreshPrevWord()

        chineseStr = self.chinese.text()
        if chineseStr == '':
            self.word.occurs -= 1
            self.lex.updateOccursOfWord(self.word)
            return

        if re.search(chineseStr, self.word.chinese, re.X) is not None:
            self.word.shots += 1
            self.lex.updateShotsOfWord(self.word)
            self.hintImage.setPixmap(self.posImage)
            self.hintImage.update()
            
        self.chinese.clear()

    def magic(self):
        self.updateWord()
        self.lex.wb.save(self.lex.filename)
        self.retrieveWord()

    
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)

    widget = MyWidget()
    widget.resize(800, 600)
    widget.show()
    
    sys.exit(app.exec_())
